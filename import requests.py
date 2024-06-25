import requests    #리퀘스트 불러주고
from bs4 import BeautifulSoup    #정리를 위해서 뷰티풀소프 불러와주고(이따가 '파서' 사용)
from openpyxl import Workbook

"""
url = 'https://cse.knu.ac.kr/bbs/board.php?bo_table=sub5_1&lang=koru'

# excel에 넣어보자고
wb = Workbook(write_only=True)
# sheet도 만들어주고
ws = wb.create_sheet('CSE_Notice')
# excel 헤더 생성
ws.append(['공지사항'])

response = requests.get(url)
notice_page = response.text

soup = BeautifulSoup(notice_page, 'html.parser')  #보기 편하게 파서로 정리    
'''
print(soup.prettify())   #보기 편한 형태로 프린트(공백도 추가)
'''
notice_title_tags = soup.select('tr td.td_subject div.bo_tit a')
print('Add Data to File...')
'''
print(notice_title_tags)  #태그별로 프린트
'''
#태그를 제외하고 텍스트만 출력
for tag in notice_title_tags:
    row = [tag.get_text().strip()]
    ws.append(row)  #태그 안에 들어있는 내용 출력하기    
wb.save('경북대_컴퓨터학부_공지사항.xlsx')   # 파일 저장
print('Save File Success!')
'''
notice_hrefs = soup.select('tr td.td_subject div.bo_tit a')

for tag in notice_hrefs:
    print(tag['href'])
'''
"""
mainUrl = 'https://cse.knu.ac.kr/bbs/board.php?bo_table=sub5_1&lang=koru'
response = requests.get(mainUrl)
mainPage = response.text
soup = BeautifulSoup(mainPage, 'html.parser')
lastPagehref = soup.select_one('a.pg_end')
lastPageind = lastPagehref['href'].find('page=') + len('page=')
lastPage = int(lastPagehref['href'][lastPageind:])
# excel 생성
wb = Workbook(write_only=True)
# sheet 생성
ws = wb.create_sheet('CSE.Notice')
# excel 헤더 생성
ws.append(['공지사항','카테고리','작성자','날짜','링크주소'])

for i in range(1,int(lastPage) + 1):
    url = f""